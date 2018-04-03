# -*- coding: utf-8 -*-

from src.instrument import Instrument
from csv import reader, writer, QUOTE_MINIMAL
from os.path import realpath
import openpyxl as oxl
import src.save as save

from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.image import Image
from kivy.properties import NumericProperty, ObjectProperty, StringProperty, DictProperty, ListProperty
from kivy.clock import Clock
from kivy.uix.listview import ListView
from kivy.uix.togglebutton import ToggleButton
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.label import Label
# _____________________________________________________________________________________________________________________
# Toolbar


class About(Popup):
    pass


class New(Popup):
    toolbar = ObjectProperty()

    # Prüfer:
    name_spinner = ObjectProperty()
    # Auftragsbezeichung:
    meas_number = StringProperty()
    # Leuchte:
    light_spinner = ObjectProperty()
    # Anzahl:
    number_light = StringProperty()

    def build(self, toolbar_parent):
        self.toolbar = toolbar_parent

        self.name_spinner.values = self.toolbar.parent.personal
        self.meas_number = self.toolbar.parent.meas_number
        self.light_spinner.values = self.toolbar.parent.leuchten["Referenznummer"]
        self.number_light = str(self.toolbar.parent.number_light)

    def make_new_file(self, name, meas_number, testing_light, number_light):
        if name != "Namen auswählen" and \
                meas_number != "" and \
                testing_light != "Leuchte auswählen" and \
                number_light != "" and number_light != "0":
            self.toolbar.parent.tester_name = name
            self.toolbar.parent.meas_number = meas_number
            self.toolbar.parent.testing_light = testing_light
            self.toolbar.parent.number_light = int(number_light)
            self.dismiss()


class Save(Popup):
    toolbar = ObjectProperty()
    save_dire = StringProperty("")

    def build(self, toolbar_parent):
        self.toolbar = toolbar_parent
        self.save_dire = self.toolbar.parent.save_dire

    def save_file_name(self, dire):
        self.toolbar.parent.save_dire = dire
        self.toolbar.parent.save_settings()
        self.dismiss()


class Instr(ListView):
    def find_all_instruments(self, instr_list):
        for name in instr_list:
            self.adapter.data.extend([name])
        self.adapter.data.remove("")
        self._trigger_reset_populate()


class Loading(Image):

    def found(self):
        self.source = "icons/Tick.png"

    def not_found(self):
        self.source = "icons/Close.png"

    def reset(self):
        self.source = "icons/b_pxl.png"


class Set(Popup):
    instr_list_view = ObjectProperty()
    load_icon_view = ObjectProperty()
    toolbar = ObjectProperty()
    device_id = StringProperty("")

    def build(self, toolbar_parent):
        self.toolbar = toolbar_parent
        # Search all instruments
        instr_list = self.toolbar.parent.instrument.show_instr_list()
        self.instr_list_view.find_all_instruments(instr_list)

    def connect_to_instrument(self):
        if self.instr_list_view.adapter.selection:
            selection = self.instr_list_view.adapter.selection[0].text
            # Connect to Selected Instrument
            if self.toolbar.parent.instrument.connect_to_device(selection):
                try:
                    self.device_id = self.toolbar.parent.instrument.get_id()
                    self.toolbar.parent.instr_name = selection
                    self.toolbar.parent.save_settings()
                    self.load_icon_view.found()
                except Exception:
                    self.device_id = ""
                    self.load_icon_view.not_found()
            else:
                self.device_id = ""
                self.load_icon_view.not_found()
        else:
            self.device_id = ""
            self.load_icon_view.not_found()


class ToolBar(GridLayout):
    mode = StringProperty("")

    def set_instrument(self):
        s = Set()
        s.open()
        s.build(self)

    def new_file(self):
        n = New()
        n.open()
        n.build(self)

    def save_file(self):
        s = Save()
        s.open()
        s.build(self)

    @staticmethod
    def about():
        About().open()


# _____________________________________________________________________________________________________________________
# Main
class MainWindow(BoxLayout):
    toolbar = ObjectProperty()
    # Settings:
    instr_name = StringProperty("")  # Save in settings
    save_dire = StringProperty("")  # Save in settings
    saved_as = StringProperty("")
    # From-Excel-Files
    personal = ListProperty()
    leuchten = DictProperty()
    # Instrument:
    instrument = ObjectProperty()
    instr_state = ListProperty([False, False, False, False])
    # Setup:
    tester_name = StringProperty()      # Name des Testers
    meas_number = StringProperty()      # Name der Messung
    testing_light = StringProperty()   # Ausgewählte Leuchte aus Excel-Datei
    number_light = NumericProperty()    # Anzahl Leuchten (int)
    curr_light = NumericProperty()      # Aktuelle Leuchte (int)
    # Measurement-Switch:
    buttons_label = ObjectProperty()
    test_widgets = DictProperty()
    # Messages:
    meas_message = StringProperty()
    meas_in_range = ObjectProperty()
    meas_in_range_label = ObjectProperty()
    # Results:
    results = DictProperty()
    io_nio = ListProperty([0, 0])

    def build(self):
        # Startvorgang des Programms:
        self.instrument = Instrument()
        # Laden der Einstellungen:
        self.get_measurement_data()
        # Verbinden zum Gerät:
        if self.instrument.connect_to_device(self.instr_name):
            print("Verbindung zum gerät hergestellt")
        else:
            print("Keine Verbindung zum Gerät möglich")
        # Laden der Excel-Dateien:
        self.load_from_excel()
        # Sequenz zur Erkennung ob Daten ausgefüllt wurden:
        Clock.schedule_interval(self.init_measurement, 1. / 10.)
        # Starten des ersten Fensters zum ausfüllen der Daten:
        Clock.schedule_once(lambda dt: self.toolbar.new_file(), 0.2)

    def get_measurement_data(self):
        print("Laden der Einstellungen aus:")
        print(realpath("src/settings.csv"))
        try:
            self.load_settings()
            print("Laden erfolgreich!")
        except Exception:
            print("Laden war nicht möglich!")
            pass
    # _________________________________________________________________________________________________________________
    # Einstellungen

    def load_settings(self):
        save_file = []
        with open(realpath("src/settings.csv"), "r", newline="") as csv_file:
            r = reader(csv_file, delimiter=" ", quotechar="|")
            for x in r:
                save_file.append("".join(x))
        # Load data
        print("Einstellungen:", save_file)
        self.save_dire, name, self.instr_name = save_file

    def save_settings(self):
        # Save data
        save_file = self.save_dire, "", self.instr_name

        with open(realpath("src/settings.csv"), "w", newline="") as csv_file:
            r = writer(csv_file, delimiter=" ", quotechar="|", quoting=QUOTE_MINIMAL)
            for x in save_file:
                r.writerow(x)

    # _________________________________________________________________________________________________________________
    # Laden aus den Excel-Dateien

    def load_from_excel(self):
        # Laden der Leuchtendaten:
        wb = oxl.load_workbook(realpath("excel_datei_einstellungen/Leuchten.xlsx"))
        wsh = wb.active
        self.leuchten["Referenznummer"] = []
        self.leuchten["Spannung"] = []
        self.leuchten["LED1Minimalstrom"] = []
        self.leuchten["LED1Maximalstrom"] = []
        self.leuchten["LED2Minimalstrom"] = []
        self.leuchten["LED2Maximalstrom"] = []

        for row in wsh.iter_rows(row_offset=1):
            if row[0].value is not None:
                self.leuchten["Referenznummer"].append(row[0].value)
                self.leuchten["Spannung"].append(row[1].value)
                self.leuchten["LED1Minimalstrom"].append(row[2].value/1000)
                self.leuchten["LED1Maximalstrom"].append(row[3].value/1000)
                self.leuchten["LED2Minimalstrom"].append(row[4].value/1000)
                self.leuchten["LED2Maximalstrom"].append(row[5].value/1000)

        # Laden der Personalladen:
        wb = oxl.load_workbook(realpath("excel_datei_einstellungen/Personal.xlsx"))
        wsh = wb.active

        for row in wsh.iter_rows(row_offset=1):
            if row[0].value is not None:
                self.personal.append(row[0].value)

        # Laden der möglichen optischen Fehler:
        wb = oxl.load_workbook(realpath("excel_datei_einstellungen/optische_Fehler.xlsx"))
        wsh = wb.active
        self.leuchten["optischeFehler"] = []

        for row in wsh.iter_rows(row_offset=1):
            if row[0].value is not None:
                self.leuchten["optischeFehler"].append(row[0].value)

        print(self.leuchten)

    # _________________________________________________________________________________________________________________
    # Messung

    def init_measurement(self, dt):
        # Falls alle Daten eingegeben und ausgewählt:
        if self.tester_name == "" or \
                self.meas_number == "" or \
                self.number_light == 0 or \
                self.testing_light == "Leuchte auswählen":
            self.meas_message = "Bitte neue Messung einrichten\n(weißes Blatt oben anklicken)."
        else:
            # Zurücksetzen alter Werte:
            self.meas_in_range.reset()
            self.meas_in_range_label.text = ""
            # Hinzufügen des Startknopfes
            self.test_widgets["start_box"] = BoxLayout(orientation="horizontal")
            self.test_widgets["start_box"].add_widget(Label())
            self.test_widgets["Messung_starten"] = ToggleButton(size_hint_y=None, height=35, text="Messung starten")
            self.test_widgets["start_box"].add_widget(self.test_widgets["Messung_starten"])
            self.test_widgets["start_box"].add_widget(Label())
            self.buttons_label.add_widget(self.test_widgets["start_box"])
            self.curr_light = 1  # Start at 1
            # Hinzufügen der Ergebnisse:
            self.io_nio = [0, 0]
            self.results["Stromwerte"] = []
            self.results["Leuchten_iO"] = []
            self.results["Fehler"] = []
            Clock.schedule_interval(self.start_measurement, 1./60.)
            Clock.unschedule(self.init_measurement)

    def start_measurement(self, dt):
        if self.instrument.connected:
            self.meas_message = "Gerät erkannt.\n" \
                                "Messung kann gestartet werden."
        if self.instrument.connected and self.test_widgets["Messung_starten"].state == "down":
            self.meas_message = "Messung wurde gestartet."
            self.buttons_label.remove_widget(self.test_widgets["start_box"])
            self.init_channel()
            Clock.unschedule(self.start_measurement)
        elif not self.instrument.connected:
            self.meas_message = "Kein Gerät eingerichtet\n(Oben links einstellen)."

    def get_testing_light_nr(self):
        return [i for i, x in enumerate(self.leuchten["Referenznummer"]) if x == self.testing_light][0]

    def init_channel(self):
        # Nummer der ausgewählten Leuchte
        nr = self.get_testing_light_nr()
        # Spannung und Strom
        volt = self.leuchten["Spannung"][nr]
        curr = "MAX"
        # Kanäle einstellen
        self.instrument.ch_set(ch_nr=0, volt=volt, curr=curr)
        self.instrument.ch_set(ch_nr=1, volt=volt, curr=curr)
        # Kanäle einschalten
        self.instrument.instr_on(0)
        self.instrument.instr_on(1)
        self.instrument.gen_on()
        # Nach etwas Zeit kann erst mit der Messung begonnen werden
        Clock.schedule_once(lambda dt: self.connect_light(), 2)

    # Anschließen der Leuchten:
    def connect_light(self):
        self.meas_message = "Bitte Anschließen der {}/{} Leuchte".format(self.curr_light, self.number_light)
        Clock.schedule_interval(self.listen_channel_connected, 0.4)

    def disconnect_light(self):
        self.meas_message = "Bitte diese Leuchte entfernen"
        Clock.schedule_interval(self.listen_channel_disconnected, 0.4)

    def listen_channel_disconnected(self, dt):
        volt1, curr1 = self.instrument.ch_measure(0)
        if curr1 < 0.001:
            Clock.schedule_once(lambda dt: self.connect_light(), 1)
            Clock.unschedule(self.listen_channel_disconnected)

    def listen_channel_connected(self, dt):
        volt1, curr1 = self.instrument.ch_measure(0)
        if curr1 > 0.001:
            self.meas_message = "Bitte warten bis sich der Strom stabilisiert hat."
            Clock.schedule_once(lambda dt: self.measure_light(), 3)
            Clock.unschedule(self.listen_channel_connected)

    def measure_light(self):
        nr = self.get_testing_light_nr()
        # Messen des Stromes beider Kanäle:
        curr = self.instrument.ch_measure(ch_nr=0)[1], self.instrument.ch_measure(ch_nr=1)[1]
        self.results["Stromwerte"].append(curr)

        led1 = self.leuchten["LED1Minimalstrom"][nr], self.leuchten["LED1Maximalstrom"][nr]
        led2 = self.leuchten["LED2Minimalstrom"][nr], self.leuchten["LED2Maximalstrom"][nr]

        # Anfragen ob im Bereich (wenn nur 0, 0 angegeben wird der Vergleich übersprungen)
        if led1 != (0, 0):
            led1_in_range = led1[0] <= curr[0] <= led1[1]
        else:
            led1_in_range = True

        if led2 != (0, 0):
            led2_in_range = led2[0] <= curr[1] <= led2[1]
        else:
            led2_in_range = True

        if led1_in_range and led2_in_range:
            # Messung liegt im Bereich
            self.results["Leuchten_iO"].append(True)
            self.meas_in_range.found()
            self.meas_in_range_label.text = "[color=#268d0d]Messwerte in Ordnung[/color]"
            self.optical_testing_init()
        else:
            # Messung liegt nicht im Bereich
            self.meas_in_range.not_found()
            self.meas_in_range_label.text = "[color=#ff0000]Messwerte liegen nicht im Bereich!" \
                                            "{}, {}, {}[/color]".format(self.results["Stromwerte"][-1],
                                                                        self.leuchten["LED1Maximalstrom"][nr],
                                                                        self.leuchten["LED2Maximalstrom"][nr])
            self.add_buttons_measurement()

    def add_buttons_measurement(self):
        # Hinzufügen der Buttons um nochmal zu messen oder fortzufahren:
        self.test_widgets["Messung_fortfahren"] = Button(size_hint_y=None, height=35, text="Messung trotzdem forfahren")
        self.test_widgets["Messung_fortfahren"].bind(on_release=self.continue_meas)
        self.test_widgets["Messung_wiederholen"] = Button(size_hint_y=None, height=35, text="Messung wiederholen")
        self.test_widgets["Messung_wiederholen"].bind(on_release=self.remeasure)
        self.test_widgets["Box_Messung"] = BoxLayout(orientation="horizontal")
        self.test_widgets["Box_Messung"].add_widget(self.test_widgets["Messung_wiederholen"])
        self.test_widgets["Box_Messung"].add_widget(self.test_widgets["Messung_fortfahren"])
        self.buttons_label.add_widget(self.test_widgets["Box_Messung"])
        self.meas_message = "[color=ff0000]Messwerte liegen nicht im Bereich![/color]"

    # Messung wiederholen
    def remeasure(self, inst):
        self.meas_message = "Messung wird wiederholt!"
        # Buttons werden entfernt:
        self.buttons_label.remove_widget(self.test_widgets["Box_Messung"])
        # Altes Messergebnis entfernen
        del self.results["Stromwerte"][-1]
        # Messvorgang wieder beginnen
        Clock.schedule_interval(self.listen_channel_connected, 0.4)

    # Messung trotzdem forfahren
    def continue_meas(self, inst):
        # Buttons werden entfernt:
        self.buttons_label.remove_widget(self.test_widgets["Box_Messung"])
        # Fortfahren
        self.results["Leuchten_iO"].append(False)
        self.optical_testing_init()

    # Optisches testen
    def add_buttons_optical_test(self):
        # Knöpfe einstellen
        self.test_widgets["Leuchte_ok"] = Button(text="Leuchte ok")
        self.test_widgets["Leuchte_ok"].bind(on_release=self.light_works)
        self.test_widgets["Leuchte_fehlerhaft"] = Button(text="Leuchte fehlerhaft")
        self.test_widgets["Leuchte_fehlerhaft"].bind(on_release=self.light_defect)

        self.test_widgets["Strom_umstellen_LED1"] = Button(text="Weiße LEDs einschalten / ausschalten")
        self.test_widgets["Strom_umstellen_LED1"].bind(on_release=self.switch_light1)

        self.test_widgets["Strom_umstellen_LED2"] = Button(text="Rote LEDs einschalten / ausschalten")
        self.test_widgets["Strom_umstellen_LED2"].bind(on_release=self.switch_light2)

        # Widget vorbereiten
        self.test_widgets["Box_optisch"] = BoxLayout(orientation="vertical")
        self.test_widgets["Box_optisch"].add_widget(Label())

        # Erste Reihe:
        box_ok = BoxLayout(orientation="horizontal", size_hint_y=None, height=35)
        box_ok.add_widget(Label())
        box_ok.add_widget(self.test_widgets["Leuchte_ok"])
        box_ok.add_widget(self.test_widgets["Leuchte_fehlerhaft"])
        box_ok.add_widget(Label())
        self.test_widgets["Box_optisch"].add_widget(box_ok)

        # Zweite Reihe:
        box_led1 = BoxLayout(orientation="horizontal", size_hint_y=None, height=35)
        box_led1.add_widget(Label())
        box_led1.add_widget(self.test_widgets["Strom_umstellen_LED1"])
        box_led1.add_widget(Label())
        self.test_widgets["Box_optisch"].add_widget(box_led1)

        # Dritte Reihe
        box_led2 = BoxLayout(orientation="horizontal", size_hint_y=None, height=35)
        box_led2.add_widget(Label())
        box_led2.add_widget(self.test_widgets["Strom_umstellen_LED2"])
        box_led2.add_widget(Label())
        self.test_widgets["Box_optisch"].add_widget(box_led2)

        # Widgets einfügen
        self.buttons_label.add_widget(self.test_widgets["Box_optisch"])
        self.meas_message = "Mit dem Schaltern die LEDs ein-/ausschalten"

    def optical_testing_init(self):
        self.add_buttons_optical_test()
        self.instrument.instr_off(0)
        self.instrument.instr_off(1)

    def switch_light1(self, inst):
        # Strom umstellen falls Knopf gedrückt wird:
        if self.instr_state[0] is False:
            self.instrument.instr_on(0)
            self.instr_state[0] = True
        else:
            self.instrument.instr_off(0)
            self.instr_state[0] = False

    def switch_light2(self, inst):
        # Strom umstellen falls Knopf gedrückt wird:
        if self.instr_state[1] is False:
            self.instrument.instr_on(1)
            self.instr_state[1] = True
        else:
            self.instrument.instr_off(1)
            self.instr_state[1] = False

    def light_defect(self, inst):
        self.buttons_label.remove_widget(self.test_widgets["Box_optisch"])
        # Falls Leuchte defekt
        # Hinzufügen von Fehler-Knöpfen zur Auswahl des Defekts
        err_box = BoxLayout(orientation="vertical", size_hint_y=None)
        self.test_widgets["Fehler_Liste"] = []
        for e in self.leuchten["optischeFehler"]:
            btn = ToggleButton(size_hint_y=None, height=35, text=e)
            self.test_widgets["Fehler_Liste"].append(btn)
            err_box.add_widget(btn)

        self.test_widgets["Auswahl"] = Button(size_hint_y=None, height=35,
                                              text="Fehler bestätigen",
                                              background_color=(1, 0, 0, 1))
        self.test_widgets["Auswahl"].bind(on_release=self.add_defect)

        self.test_widgets["Box_Error"] = BoxLayout(orientation="horizontal")
        self.test_widgets["Box_Error"].add_widget(err_box)
        self.test_widgets["Box_Error"].add_widget(self.test_widgets["Auswahl"])

        self.buttons_label.add_widget(self.test_widgets["Box_Error"])

    def add_defect(self, inst):
        chosen = []
        for defect in self.test_widgets["Fehler_Liste"]:
            if defect.state == "down":
                chosen.append(defect.text)
        # Nur forfahren falls mindestens ein Defekt ausgewählt wurde
        if len(chosen) > 0:
            self.buttons_label.remove_widget(self.test_widgets["Box_Error"])
            self.results["Fehler"].append(chosen)
            # Ändern des letzten Ergebnisses falls optisch die Messung nicht in Ordnung ist
            self.results["Leuchten_iO"][-1] = False
            self.end_measurement()

    def light_works(self, inst):
        self.buttons_label.remove_widget(self.test_widgets["Box_optisch"])
        self.results["Fehler"].append([])
        self.end_measurement()

    def end_measurement(self):
        self.curr_light += 1

        liste = self.results["Leuchten_iO"]
        self.io_nio = sum(liste), len(liste) - sum(liste)

        print("EINSTELLUNGEN:", self.leuchten)
        print("ERGEBNISSE: ", self.results)

        if self.curr_light > self.number_light:
            # Kanäle ausschalten
            self.instrument.instr_off(0)
            self.instrument.instr_off(1)
            # Messergebnisse speichern
            self.save_result()
            # Messung zu Ende neues Fesnster Öffnen
            self.tester_name = ""
            self.curr_light = 1
            # Sequenz zur Erkennung ob Daten ausgefüllt wurden:
            Clock.schedule_interval(self.init_measurement, 1. / 10.)
            # Starten des ersten Fensters zum ausfüllen der Daten:
            Clock.schedule_once(lambda dt: self.toolbar.new_file(), 0.2)
        else:
            self.instrument.instr_on(0)
            self.instrument.instr_on(1)
            self.disconnect_light()

    # Speichern aller Daten
    def save_result(self):
        nr = self.get_testing_light_nr()

        data = {"tester_name": self.tester_name,
                "meas_number": self.meas_number,
                "testing_light": self.testing_light,
                "number_light": self.number_light,
                "optischeFehler": self.leuchten["optischeFehler"],
                "Spannung": self.leuchten["Spannung"][nr],
                "Strombereich_LED1": (self.leuchten["LED1Minimalstrom"][nr], self.leuchten["LED1Maximalstrom"][nr]),
                "Strombereich_LED2": (self.leuchten["LED2Minimalstrom"][nr], self.leuchten["LED2Maximalstrom"][nr])}

        results = {"Stromwerte": self.results["Stromwerte"],
                   "opt_Fehler": self.results["Fehler"]}

        save.save_as(data, results, usb_path=self.save_dire)


class MeasurementApp(App):
    icon = './icons/icon.png'

    def build(self):
        w = MainWindow()
        w.build()
        return w
